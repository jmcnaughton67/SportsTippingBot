from flask import Flask, render_template, request
import openpyxl as xl
import os
import requests
import pandas as pd
from datetime import datetime, timedelta

app = Flask(__name__)

def getFixtures(round):
    # init game count
    game_count = 0

    # Get Fixture Data
    url = 'https://www.nrl.com/draw//data'
    params = {'competition': '111', 'round': str(round), 'season': '2023'}
    response = requests.get(url, params=params)
    data = response.json()
    fixtures = []

    # Get Ladder Position data
    ladder_url = 'https://www.nrl.com/ladder/data'
    ladder_params = {'competition': '111', 'Seasons': '2023'}
    ladder_response = requests.get(ladder_url, params=ladder_params)
    ladder_data = ladder_response.json()
    ladder = []
    for team in ladder_data['positions']:
        ladder_team = team['teamNickname']
        ladder.append(ladder_team)

    # Set the fixture information for each game of the round
    for fixture in data['fixtures']:
        # Get Game information
        game_count = game_count + 1
        home_team = fixture['homeTeam']['nickName']
        home_image = 'https://www.nrl.com/.theme/' + home_team.replace(" ", "-").lower() + '/badge.svg?bust=202302142313'
        away_team = fixture['awayTeam']['nickName']
        away_image = 'https://www.nrl.com/.theme/' + away_team.replace(" ", "-").lower() + '/badge.svg?bust=202302142313'
        venue = fixture['venue']
        match_status = fixture['matchState']
        home_pos = ladder.index(home_team) + 1
        away_pos = ladder.index(away_team) + 1

        #If the game has not started, scores do not exist
        try:
            home_score = fixture['homeTeam']['score']
            away_score = fixture['awayTeam']['score']
        except:
            home_score = "0"
            away_score = "0"

        # If the game has already started - odds are no longer available, so set this to N/A in this case:
        # This will be fixed
        if match_status == "Upcoming":
            home_odds = fixture['homeTeam']['odds']
            away_odds = fixture['awayTeam']['odds']
        else:
            home_odds = "N/A"
            away_odds = "N/A"

        # parse the datetime string into a datetime object
        dt = datetime.strptime(fixture['clock']['kickOffTimeLong'], '%Y-%m-%dT%H:%M:%SZ')

        # extract the date and time components into separate variables
        date_component = dt.date()
        date = date_component.strftime('%A, %d' + ('th' if 11<=date_component.day<=13 else {1:'st', 2:'nd', 3:'rd'}.get(date_component.day%10, 'th')) + ' %B')
        time = dt - timedelta(hours=13)
        time = time.strftime('%I:%M %p')

        #Add calculation based on odds and ladder position. Return True for the most important game.
        # just using game 2 for testing until the ladder position issue is resolved
        if game_count == 2:
            isMainGame = True
        else:
            isMainGame = False

        fixtures.append({'Round': round, 'Home Team': home_team, 'Home Odds': home_odds, 'Home Position': home_pos, 'Home Image': home_image, 'Home Score': home_score, 'Away Team': away_team, 'Away Odds': away_odds, 'Away Position': away_pos,  'Away Image': away_image, 'Away Score': away_score, 'Venue': venue, 'Date': date, 'Time': time, 'MainGame': isMainGame, 'Match Status': match_status})

    df = pd.DataFrame(fixtures)
    df.index += 1
    #sort df here by MainGame, so that the game with True is at the top so it appears top
    # We don't want mainGame top of the webpage - keep it in Chronological order - JM

    return df


def createCode(code):

    filepath = "./" + code + ".xlsx"

    if os.path.isfile(filepath):
        return False
    else:
        wb = xl.Workbook()

        tip_sheet = wb.active

        tip_sheet['A1'].value = "Name:"
        tip_sheet['B1'].value = "Code:"

        wb.save(filepath)

        return True

def saveToSheet(name, code, tips, margins):

    filepath = "./" + code + ".xlsx"
    if os.path.isfile(filepath):
        wb = xl.load_workbook(filepath)
    else:
        print("Code Does Not Exist")

    tip_sheet = wb.active
    end = tip_sheet.max_row + 1
    print(end)
    for i in range(0, len(tips)):
        cell = end + 1
        tip_sheet['A'+str(cell)].value = name
        tip_sheet['B'+str(cell)].value = str(i+1)
        tip_sheet['C'+str(cell)].value = tips[i]
        tip_sheet['D'+str(cell)].value = margins[i]

    wb.save(filepath)
    print("saved to " + filepath)


def getFromSheet(code):

    filepath = "./" + code + ".xlsx"
    if os.path.isfile(filepath):
        wb = xl.load_workbook(filepath)
    else:
        return "File doesn't exist you dingbat"

    tip_sheet = wb.active
    end = str(tip_sheet.max_row)
    print(end)
    name = tip_sheet['A'+end].value
    code = tip_sheet['B'+end].value

    return name, code


@app.route('/')
def index():

    return render_template('index.html')

@app.route('/create')
def create():
    return render_template('create.html')

@app.route('/login', methods=['POST'])
def login():
    name = request.form.get('name')
    code = request.form.get('code')

    df = getFixtures(1)
    #saveToSheet(name, code)

    #name2, code2 = getFromSheet(code)

    return render_template('tipsheet.html', name=name, code=code, data=df)

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    games = request.form['games']
    code = request.form.get('code')

    tips = []
    margins = []

    for i in range(1, int(games)+1):
        tips.append(request.form['game-'+str(i)])
        margins.append(request.form['game-'+str(i)+"-margin"])

    saveToSheet(name, code, tips, margins)

    #name2, code2 = getFromSheet(code)

    return render_template('submit.html', name=name, code=code)


@app.route('/newCode', methods=['POST'])
def newCode():
    name = request.form.get('name')
    code = request.form.get('code')

    createCode(code)

    return render_template('index.html', generated ="code generated:"+code)


if __name__ == '__main__':
    app.run()

